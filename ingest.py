#!/usr/bin/env python3
"""
Ingest tentative rulings and maintain tentatives.parquet + tentatives.db.

Usage:
    python ingest.py                         # rebuild from all source files
    python ingest.py path/to/new.xlsx        # append a new export
    python ingest.py path/to/new.json        # append a new json export

tentatives.parquet  — canonical dataset, committed to git (~10 MB)
tentatives.db       — local SQLite for querying, gitignored (~100 MB)
"""

import sys
import json
import sqlite3
import hashlib
from pathlib import Path
from datetime import datetime, date, time

try:
    import pandas as pd
except ImportError:
    sys.exit("Missing dependency: pip install pandas pyarrow openpyxl")

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

HERE = Path(__file__).parent
PARQUET = HERE / "tentatives.parquet"
DB_PATH = HERE / "tentatives.db"

COLUMNS = ["case_number", "case_title", "court_date", "hearing_time",
           "calendar_matter", "judge", "ruling", "row_hash"]

def make_hash(case_number, court_date, ruling):
    key = f"{case_number}|{court_date}|{ruling}"
    return hashlib.sha1(key.encode()).hexdigest()

def normalize_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        for fmt in ("%b-%d-%Y %I:%M %p", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date().isoformat()
            except ValueError:
                continue
        try:
            return datetime.strptime(val.strip().split()[0], "%b-%d-%Y").date().isoformat()
        except ValueError:
            return val
    return str(val)

def normalize_time(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    return str(val)

def load_xlsm_2014_2018(path):
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        case_number, case_title, court_date, calendar_matter, ruling = r[:5]
        if not any([case_number, ruling]):
            continue
        hearing_time = None
        if isinstance(court_date, datetime) and (court_date.hour or court_date.minute):
            hearing_time = court_date.strftime("%H:%M")
        rows.append({
            "case_number":     str(case_number).strip() if case_number else None,
            "case_title":      str(case_title).strip() if case_title else None,
            "court_date":      normalize_date(court_date),
            "hearing_time":    hearing_time,
            "calendar_matter": str(calendar_matter).strip() if calendar_matter else None,
            "judge":           None,
            "ruling":          str(ruling).strip() if ruling else None,
        })
    return rows

def load_xlsx_2020_plus(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["tentatives"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        case_number, case_title, court_date, hearing_time, calendar_matter, judge, ruling = r[:7]
        if not any([case_number, ruling]):
            continue
        rows.append({
            "case_number":     str(case_number).strip() if case_number else None,
            "case_title":      str(case_title).strip() if case_title else None,
            "court_date":      normalize_date(court_date),
            "hearing_time":    normalize_time(hearing_time),
            "calendar_matter": str(calendar_matter).strip() if calendar_matter else None,
            "judge":           str(judge).strip() if judge else None,
            "ruling":          str(ruling).strip() if ruling else None,
        })
    return rows

def load_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    rows = []
    for rec in data:
        court_date_raw = rec.get("Court Date", "")
        hearing_time = None
        parts = court_date_raw.strip().split()
        if len(parts) >= 3:
            try:
                t = datetime.strptime(f"{parts[1]} {parts[2]}", "%I:%M %p")
                hearing_time = t.strftime("%H:%M")
            except ValueError:
                pass
        rows.append({
            "case_number":     rec.get("Case Number", "").strip() or None,
            "case_title":      rec.get("Case Title", "").strip() or None,
            "court_date":      normalize_date(court_date_raw),
            "hearing_time":    hearing_time,
            "calendar_matter": rec.get("Calendar Matter", "").strip() or None,
            "judge":           None,
            "ruling":          rec.get("Rulings", "").strip() or None,
        })
    return rows

def detect_and_load(path):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_json(path)
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, keep_vba=(suffix == ".xlsm"))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        if "Judge" in headers or "Hearing Time" in headers:
            return load_xlsx_2020_plus(path)
        return load_xlsm_2014_2018(path)
    raise ValueError(f"Unsupported file type: {path}")

def to_df(rows):
    df = pd.DataFrame(rows, columns=COLUMNS)
    df["row_hash"] = df.apply(
        lambda r: make_hash(r["case_number"] or "", r["court_date"] or "", r["ruling"] or ""), axis=1
    )
    return df

def merge(existing: pd.DataFrame, new: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    combined = pd.concat([existing, new], ignore_index=True)
    before = len(combined)
    combined = combined.drop_duplicates(subset="row_hash", keep="first")
    after = len(combined)
    inserted = after - len(existing)
    skipped = before - after - max(0, len(new) - (before - len(existing)))
    return combined.sort_values("court_date").reset_index(drop=True), inserted, len(new) - inserted

def save_parquet(df: pd.DataFrame):
    df.to_parquet(PARQUET, index=False, compression="zstd")

def save_sqlite(df: pd.DataFrame):
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        DROP TABLE IF EXISTS tentatives;
        CREATE TABLE tentatives (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            case_number     TEXT,
            case_title      TEXT,
            court_date      DATE,
            hearing_time    TEXT,
            calendar_matter TEXT,
            judge           TEXT,
            ruling          TEXT,
            row_hash        TEXT UNIQUE
        );
        CREATE INDEX IF NOT EXISTS idx_case_number ON tentatives(case_number);
        CREATE INDEX IF NOT EXISTS idx_court_date  ON tentatives(court_date);
        CREATE INDEX IF NOT EXISTS idx_judge       ON tentatives(judge);
    """)
    df.to_sql("tentatives", conn, if_exists="append", index=False,
              method="multi", chunksize=500)
    conn.close()

def summary(df: pd.DataFrame):
    print(f"\nDataset: {len(df):,} rows  |  {df['court_date'].min()} → {df['court_date'].max()}")
    print(f"Parquet: {PARQUET.stat().st_size / 1e6:.1f} MB   DB: {DB_PATH.stat().st_size / 1e6:.1f} MB")

def main():
    existing = pd.read_parquet(PARQUET) if PARQUET.exists() else pd.DataFrame(columns=COLUMNS)

    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            p = Path(arg)
            if not p.exists():
                print(f"Not found: {p}")
                continue
            print(f"Loading {p.name}...")
            new_df = to_df(detect_and_load(p))
            existing, inserted, skipped = merge(existing, new_df)
            print(f"  {inserted} inserted, {skipped} skipped (duplicates)")
    else:
        sources = [
            HERE / "tentatives.xlsm",
            HERE / "sfsc tentatives 01-2020 to 07-2025.xlsx",
        ]
        for src in sources:
            if not src.exists():
                print(f"Not found: {src.name}")
                continue
            print(f"Loading {src.name}...")
            new_df = to_df(detect_and_load(src))
            existing, inserted, skipped = merge(existing, new_df)
            print(f"  {inserted} inserted, {skipped} skipped (duplicates)")

    print("Saving parquet...")
    save_parquet(existing)
    print("Saving SQLite...")
    save_sqlite(existing)
    summary(existing)

if __name__ == "__main__":
    main()
